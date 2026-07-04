from __future__ import annotations

import io
import math
import re
import statistics
from collections import defaultdict
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


COUNTRY_SHEET_EXCLUDE = {"pricing data"}
ERROR_PREFIXES = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def parse_mhr_workbook(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """Parse an MHR calculator workbook into dashboard-ready analytics."""
    source = io.BytesIO(file_bytes)
    formula_workbook = openpyxl.load_workbook(source, data_only=False)

    source.seek(0)
    value_workbook = openpyxl.load_workbook(source, data_only=True)

    pricing_data = _parse_pricing_data(formula_workbook, value_workbook)
    country_sheets = _detect_country_sheets(formula_workbook)
    records: list[dict[str, Any]] = []
    quality = {
        "formula_errors": [],
        "warnings": [],
        "missing_clean_room_cost": 0,
        "blank_mhr": 0,
    }

    for sheet_name in country_sheets:
        formula_sheet = formula_workbook[sheet_name]
        value_sheet = value_workbook[sheet_name]
        records.extend(_parse_country_sheet(formula_sheet, value_sheet, quality))
        _scan_formula_warnings(formula_sheet, quality)

    assumptions = _pricing_assumptions(pricing_data)
    summary = _build_summary(records, pricing_data, quality, filename, formula_workbook.sheetnames)

    return {
        "filename": filename,
        "sheets": formula_workbook.sheetnames,
        "regions": country_sheets,
        "summary": summary,
        "records": records,
        "regionSummary": _region_summary(records),
        "processSummary": _process_summary(records),
        "machineComparison": _machine_comparison(records),
        "assumptions": assumptions,
        "quality": _quality_report(records, quality, pricing_data, country_sheets),
    }


def _detect_country_sheets(workbook: openpyxl.Workbook) -> list[str]:
    sheets = []
    for worksheet in workbook.worksheets:
        if worksheet.title.lower() in COUNTRY_SHEET_EXCLUDE:
            continue
        if worksheet["A1"].value and worksheet["D1"].value == "MHR":
            sheets.append(worksheet.title)
    return sheets


def _parse_country_sheet(formula_sheet, value_sheet, quality: dict[str, Any]) -> list[dict[str, Any]]:
    headers = _headers(formula_sheet)
    lookup = _header_lookup(headers)
    region = _clean(value_sheet["C3"].value or formula_sheet.title)
    last_row = _last_machine_row(value_sheet)
    rows = []

    for row_idx in range(3, last_row + 1):
        machine = _clean(value_sheet.cell(row_idx, 1).value)
        if not machine:
            continue

        mhr = _number(value_sheet.cell(row_idx, 4).value)
        if mhr is None:
            quality["blank_mhr"] += 1

        details = _clean(value_sheet.cell(row_idx, 2).value)
        process = _extract_process(details, region, formula_sheet.cell(row_idx, 2).value)
        clean_room_cost = _value_by_key(value_sheet, row_idx, lookup, "clean_room_cost")
        if clean_room_cost in (None, ""):
            quality["missing_clean_room_cost"] += 1

        rows.append(
            {
                "row": row_idx,
                "region": region,
                "machine": machine,
                "details": details,
                "process": process,
                "mhr": mhr,
                "setupTimeMinutes": _value_by_key(value_sheet, row_idx, lookup, "setup_time"),
                "weeksPerYear": _value_by_key(value_sheet, row_idx, lookup, "weeks"),
                "daysPerWeek": _value_by_key(value_sheet, row_idx, lookup, "days"),
                "shiftsPerDay": _value_by_key(value_sheet, row_idx, lookup, "shifts"),
                "hoursPerShift": _value_by_key(value_sheet, row_idx, lookup, "hours"),
                "totalMachineTime": _value_by_key(value_sheet, row_idx, lookup, "total_machine_time"),
                "oee": _value_by_key(value_sheet, row_idx, lookup, "oee"),
                "netMachineTime": _value_by_key(value_sheet, row_idx, lookup, "net_machine_time"),
                "capitalCostInput": _value_by_key(value_sheet, row_idx, lookup, "capital_cost_input"),
                "depreciationYears": _value_by_key(value_sheet, row_idx, lookup, "depreciation"),
                "interestRate": _value_by_key(value_sheet, row_idx, lookup, "interest_rate"),
                "annualCapitalCost": _value_by_key(value_sheet, row_idx, lookup, "annual_capital_cost"),
                "floorSpace": _value_by_key(value_sheet, row_idx, lookup, "floor_space"),
                "floorRent": _value_by_key(value_sheet, row_idx, lookup, "floor_rent"),
                "fixedCost": _value_by_key(value_sheet, row_idx, lookup, "fixed_cost"),
                "powerConsumption": _value_by_key(value_sheet, row_idx, lookup, "power"),
                "energyCost": _value_by_key(value_sheet, row_idx, lookup, "energy_cost"),
                "consumables": _value_by_key(value_sheet, row_idx, lookup, "consumables"),
                "maintenance": _value_by_key(value_sheet, row_idx, lookup, "maintenance"),
                "variableCost": _value_by_key(value_sheet, row_idx, lookup, "variable_cost"),
                "machineHourWithoutIndirect": _value_by_key(value_sheet, row_idx, lookup, "machine_wo_indirect"),
                "overhead": _value_by_key(value_sheet, row_idx, lookup, "overhead"),
                "machineHourWithoutLabor": _value_by_key(value_sheet, row_idx, lookup, "machine_without_labor"),
                "spm": _value_by_key(value_sheet, row_idx, lookup, "spm"),
                "cleanRoomCost": clean_room_cost,
            }
        )

        for col_idx in range(1, formula_sheet.max_column + 1):
            cached = value_sheet.cell(row_idx, col_idx).value
            if isinstance(cached, str) and cached.startswith(ERROR_PREFIXES):
                quality["formula_errors"].append(
                    {
                        "sheet": formula_sheet.title,
                        "cell": f"{get_column_letter(col_idx)}{row_idx}",
                        "value": cached,
                    }
                )

    return rows


def _headers(sheet) -> dict[int, str]:
    return {idx: _clean(sheet.cell(1, idx).value) for idx in range(1, sheet.max_column + 1)}


def _header_lookup(headers: dict[int, str]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for col_idx, header in headers.items():
        text = header.lower()
        if text == "setup time":
            lookup["setup_time"] = col_idx
        elif text == "# of weeks":
            lookup["weeks"] = col_idx
        elif text == "# of days":
            lookup["days"] = col_idx
        elif text == "# of shifts":
            lookup["shifts"] = col_idx
        elif text == "# of hours":
            lookup["hours"] = col_idx
        elif text == "total machine time":
            lookup["total_machine_time"] = col_idx
        elif "efficiency" in text:
            lookup["oee"] = col_idx
        elif text == "net machine time":
            lookup["net_machine_time"] = col_idx
        elif text == "capital costs incl.":
            lookup["capital_cost_input"] = col_idx
        elif text == "depreciation time":
            lookup["depreciation"] = col_idx
        elif text == "annual interest rate":
            lookup["interest_rate"] = col_idx
        elif text == "capital cost":
            lookup["annual_capital_cost"] = col_idx
        elif text == "floor space used":
            lookup["floor_space"] = col_idx
        elif text == "floor rent":
            lookup["floor_rent"] = col_idx
        elif text == "fix costs":
            lookup["fixed_cost"] = col_idx
        elif text == "power consumption @ peak load":
            lookup["power"] = col_idx
        elif text == "energy costs":
            lookup["energy_cost"] = col_idx
        elif text.startswith("yearly consumables"):
            lookup["consumables"] = col_idx
        elif text.startswith("yearly equipment maintenance"):
            lookup["maintenance"] = col_idx
        elif text == "variable cost":
            lookup["variable_cost"] = col_idx
        elif text.startswith("m/c hour wo indirect"):
            lookup["machine_wo_indirect"] = col_idx
        elif text.startswith("additional overhead"):
            lookup["overhead"] = col_idx
        elif text.startswith("m/c hour without labor"):
            lookup["machine_without_labor"] = col_idx
        elif text == "spm":
            lookup["spm"] = col_idx
        elif text == "clean room cost":
            lookup["clean_room_cost"] = col_idx
    return lookup


def _value_by_key(sheet, row_idx: int, lookup: dict[str, int], key: str) -> Any:
    col_idx = lookup.get(key)
    if not col_idx:
        return None
    value = sheet.cell(row_idx, col_idx).value
    return _number(value) if isinstance(value, (int, float)) else value


def _last_machine_row(sheet) -> int:
    last = 2
    for row_idx in range(3, sheet.max_row + 1):
        if sheet.cell(row_idx, 1).value is not None:
            last = row_idx
    return last


def _extract_process(details: str, region: str, formula_value: Any) -> str:
    if details and region and details.upper().endswith(f"-{region.upper()}"):
        return details[: -(len(region) + 1)].strip()
    if isinstance(formula_value, str):
        match = re.search(r'CONCATENATE\(" ?([^"]+)"', formula_value)
        if match:
            return match.group(1).strip()
    if "-" in details:
        return details.rsplit("-", 1)[0].strip()
    return details or "Unknown"


def _parse_pricing_data(formula_workbook, value_workbook) -> dict[str, dict[str, Any]]:
    if "Pricing Data" not in formula_workbook.sheetnames:
        return {}

    formula_sheet = formula_workbook["Pricing Data"]
    value_sheet = value_workbook["Pricing Data"]
    countries = []
    for col_idx in range(3, formula_sheet.max_column + 1):
        country = _clean(value_sheet.cell(6, col_idx).value)
        if country:
            countries.append((country, col_idx))

    pricing: dict[str, dict[str, Any]] = {}
    for country, col_idx in countries:
        pricing[country] = {}
        for row_idx in range(7, 21):
            metric = _clean(value_sheet.cell(row_idx, 2).value)
            if not metric:
                continue
            pricing[country][metric] = value_sheet.cell(row_idx, col_idx).value
    return pricing


def _pricing_assumptions(pricing: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for country, metrics in pricing.items():
        item = {"region": country}
        item.update(metrics)
        rows.append(item)
    return rows


def _build_summary(
    records: list[dict[str, Any]],
    pricing_data: dict[str, dict[str, Any]],
    quality: dict[str, Any],
    filename: str,
    sheets: list[str],
) -> dict[str, Any]:
    mhrs = [_number(row["mhr"]) for row in records if _number(row["mhr"]) is not None]
    regions = sorted({row["region"] for row in records})
    processes = sorted({row["process"] for row in records})
    machines = sorted({row["machine"] for row in records})
    cheapest_region = min(_region_summary(records), key=lambda row: row["averageMhr"], default=None)
    expensive_region = max(_region_summary(records), key=lambda row: row["averageMhr"], default=None)

    return {
        "filename": filename,
        "sheetCount": len(sheets),
        "recordCount": len(records),
        "machineCount": len(machines),
        "regionCount": len(regions),
        "processCount": len(processes),
        "averageMhr": _round(statistics.mean(mhrs)) if mhrs else None,
        "medianMhr": _round(statistics.median(mhrs)) if mhrs else None,
        "minMhr": _round(min(mhrs)) if mhrs else None,
        "maxMhr": _round(max(mhrs)) if mhrs else None,
        "cheapestRegion": cheapest_region["region"] if cheapest_region else None,
        "mostExpensiveRegion": expensive_region["region"] if expensive_region else None,
        "pricingRegionCount": len(pricing_data),
        "formulaErrorCount": len(quality["formula_errors"]),
    }


def _region_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups = _group(records, "region")
    rows = []
    for region, items in groups.items():
        mhrs = [_number(item["mhr"]) for item in items if _number(item["mhr"]) is not None]
        if not mhrs:
            continue
        rows.append(
            {
                "region": region,
                "machines": len({item["machine"] for item in items}),
                "records": len(items),
                "averageMhr": _round(statistics.mean(mhrs)),
                "medianMhr": _round(statistics.median(mhrs)),
                "minMhr": _round(min(mhrs)),
                "maxMhr": _round(max(mhrs)),
                "averageEnergyCost": _avg(items, "energyCost"),
                "averageFloorRent": _avg(items, "floorRent"),
                "averageNetMachineTime": _avg(items, "netMachineTime"),
            }
        )
    return sorted(rows, key=lambda item: item["averageMhr"])


def _process_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups = _group(records, "process")
    rows = []
    for process, items in groups.items():
        mhrs = [_number(item["mhr"]) for item in items if _number(item["mhr"]) is not None]
        if not mhrs:
            continue
        by_region = {}
        for region, region_items in _group(items, "region").items():
            by_region[region] = _avg(region_items, "mhr")
        rows.append(
            {
                "process": process,
                "records": len(items),
                "machineCount": len({item["machine"] for item in items}),
                "averageMhr": _round(statistics.mean(mhrs)),
                "minMhr": _round(min(mhrs)),
                "maxMhr": _round(max(mhrs)),
                "byRegion": by_region,
            }
        )
    return sorted(rows, key=lambda item: item["records"], reverse=True)


def _machine_comparison(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for machine, items in _group(records, "machine").items():
        comparable = [item for item in items if _number(item["mhr"]) is not None]
        if len(comparable) < 2:
            continue
        cheapest = min(comparable, key=lambda item: item["mhr"])
        expensive = max(comparable, key=lambda item: item["mhr"])
        region_rates = {item["region"]: _round(item["mhr"]) for item in comparable}
        rows.append(
            {
                "machine": machine,
                "process": comparable[0]["process"],
                "regions": region_rates,
                "cheapestRegion": cheapest["region"],
                "cheapestMhr": _round(cheapest["mhr"]),
                "mostExpensiveRegion": expensive["region"],
                "mostExpensiveMhr": _round(expensive["mhr"]),
                "spread": _round(expensive["mhr"] - cheapest["mhr"]),
                "spreadPercent": _round((expensive["mhr"] / cheapest["mhr"] - 1) * 100) if cheapest["mhr"] else None,
            }
        )
    return sorted(rows, key=lambda item: item["spread"], reverse=True)


def _quality_report(
    records: list[dict[str, Any]],
    quality: dict[str, Any],
    pricing_data: dict[str, dict[str, Any]],
    country_sheets: list[str],
) -> dict[str, Any]:
    pricing_regions = sorted(pricing_data.keys())
    sheet_regions = sorted({record["region"] for record in records})
    missing_tabs = [region for region in pricing_regions if region not in sheet_regions]

    warnings = list(quality["warnings"])
    if missing_tabs:
        warnings.append(
            {
                "type": "missing_country_tabs",
                "severity": "info",
                "message": "Pricing Data contains regions without calculation tabs.",
                "detail": ", ".join(missing_tabs),
            }
        )
    if quality["missing_clean_room_cost"]:
        warnings.append(
            {
                "type": "missing_clean_room_cost",
                "severity": "info",
                "message": "Clean Room Cost is blank on many rows, so final MHR usually equals machine-hour-without-labor.",
                "detail": f"{quality['missing_clean_room_cost']} rows",
            }
        )
    if pricing_data and any("HIGH SKILLED LABOUR " in metrics for metrics in pricing_data.values()):
        warnings.append(
            {
                "type": "unused_labor_assumptions",
                "severity": "review",
                "message": "Labor assumptions exist in Pricing Data, but the detected MHR formula chain does not add direct labor.",
                "detail": "High-skilled and semi-skilled labor rates appear to be reference assumptions.",
            }
        )

    return {
        "warnings": warnings,
        "formulaErrors": quality["formula_errors"][:200],
        "formulaErrorCount": len(quality["formula_errors"]),
        "blankMhrCount": quality["blank_mhr"],
        "countrySheets": country_sheets,
        "pricingRegions": pricing_regions,
    }


def _scan_formula_warnings(sheet, quality: dict[str, Any]) -> None:
    found_weld_condition = False
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str) or not value.startswith("="):
                continue
            normalized = value.replace(" ", "").upper()
            if 'IF(B' in normalized and '="WELD"' in normalized:
                found_weld_condition = True
    if found_weld_condition:
        quality["warnings"].append(
            {
                "type": "weld_consumables_formula",
                "severity": "review",
                "message": f"{sheet.title} uses an exact Weld check in consumables formulas.",
                "detail": 'Rows whose Details value is like "Spot Weld-USA" may not match IF(B="Weld", ...).',
            }
        )


def _group(records: list[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[_clean(record.get(key)) or "Unknown"].append(record)
    return dict(grouped)


def _avg(items: list[dict[str, Any]], key: str) -> float | None:
    values = [_number(item.get(key)) for item in items if _number(item.get(key)) is not None]
    return _round(statistics.mean(values)) if values else None


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(value):
        return float(value)
    return None


def _round(value: float | None, digits: int = 2) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)
